import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font

INPUT = "coverity_report.xlsx"
OUTPUT = "coverity_dashboard_optA.xlsx"

# -------------------------------------------------------
# Beautify Category (does not affect detail dashboard)
# -------------------------------------------------------
def beautify(cat: str):
    if not isinstance(cat, str):
        return cat
    return " ".join(w.capitalize() for w in str(cat).split())

# -------------------------------------------------------
# Load input file
# -------------------------------------------------------
df_summary = pd.read_excel(INPUT, sheet_name="Summary")
df_summary["Category"] = df_summary["Category"].apply(beautiful := beautify)

wb = load_workbook(INPUT)

# Remove old dashboards
for s in list(wb.sheetnames):
    if s.startswith("Dashboard"):
        del wb[s]

# -------------------------------------------------------
# MAIN DASHBOARD — TWO PIE CHARTS (FIRST & LAST)
# -------------------------------------------------------
ws = wb.create_sheet("Dashboard_Main")

df_old = df_summary[df_summary["Snapshot"] == "first"].groupby("Category")["Count"].sum().reset_index()
df_last = df_summary[df_summary["Snapshot"] == "last"].groupby("Category")["Count"].sum().reset_index()

# ------ MAIN TITLE OUTSIDE CHART (to avoid broken font) ------
ws["A1"] = "TOTAL CATEGORY DISTRIBUTION (ALL PROJECTS)"
ws["A1"].font = Font(bold=True, size=16)

# --- Write OLD table ---
ws.append(["Category (FIRST)", "Count"])
for row in df_old.itertuples(index=False):
    ws.append(list(row))

# --- PIE CHART FIRST ---
pie1 = PieChart()
labels = Reference(ws, min_col=1, min_row=2, max_row=1 + len(df_old))
data = Reference(ws, min_col=2, min_row=1, max_row=1 + len(df_old))
pie1.add_data(data, titles_from_data=True)
pie1.set_categories(labels)

# Do NOT set chart.title → prevents font distortion!
ws.add_chart(pie1, "E3")

# --- Write LAST table ---
start = len(df_old) + 5
ws.cell(row=start - 1, column=1, value="Category (LAST)").font = Font(bold=True)
ws.cell(row=start - 1, column=2, value="Count").font = Font(bold=True)

for row in df_last.itertuples(index=False):
    ws.append(list(row))

# --- PIE CHART LAST ---
pie2 = PieChart()
labels2 = Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(df_last))
data2 = Reference(ws, min_col=2, min_row=start, max_row=start + len(df_last))
pie2.add_data(data2, titles_from_data=True)
pie2.set_categories(labels2)

ws.add_chart(pie2, "E20")

# -------------------------------------------------------
# DETAIL DASHBOARDS — ONLY TOTAL FIRST vs LAST
# -------------------------------------------------------
projects = df_summary["Project"].unique()

for proj in projects:
    ws2 = wb.create_sheet(f"Dashboard_{proj}")

    df_p = df_summary[df_summary["Project"] == proj]

    total_first = int(df_p[df_p["Snapshot"] == "first"]["Count"].sum())
    total_last = int(df_p[df_p["Snapshot"] == "last"]["Count"].sum())
    delta_total = total_last - total_first

    # ====== TITLE OUTSIDE CHART (fix font issue) ======
    ws2["A1"] = f"Project: {proj} – Defect Summary"
    ws2["A1"].font = Font(bold=True, size=16)

    # KPI
    ws2["A3"] = "Total FIRST"
    ws2["B3"] = total_first
    ws2["A4"] = "Total LAST"
    ws2["B4"] = total_last
    ws2["A5"] = "Delta"
    ws2["B5"] = delta_total

    # Table for chart
    ws2.append([])
    ws2.append(["Snapshot", "Count"])
    ws2.append(["first", total_first])
    ws2.append(["last", total_last])

    # BAR CHART
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.y_axis.title = "Total Defects"
    bar.x_axis.title = "Snapshot"

    data_ref = Reference(ws2, min_col=2, min_row=8, max_row=9)
    cats_ref = Reference(ws2, min_col=1, min_row=8, max_row=9)

    bar.add_data(data_ref, titles_from_data=False)
    bar.set_categories(cats_ref)

    # Show values
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True

    # DO NOT SET bar.title → avoid font bug

    ws2.add_chart(bar, "E2")

# -------------------------------------------------------
# SAVE
# -------------------------------------------------------
wb.save(OUTPUT)
print("DONE →", OUTPUT)
