import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart, RadarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

INPUT = "coverity_report.xlsx"
OUTPUT = "coverity_dashboard_optA.xlsx"

# Load summary
df_summary = pd.read_excel(INPUT, sheet_name="Summary")

# Beautify category names
def beautify(cat: str):
    if not isinstance(cat, str):
        return cat
    return " ".join(w.capitalize() for w in str(cat).split())

df_summary["Category"] = df_summary["Category"].apply(beautify)

# Copy workbook
wb = load_workbook(INPUT)

# Remove old dashboard sheets
for s in list(wb.sheetnames):
    if s.startswith("Dashboard"):
        del wb[s]

# --- MAIN DASHBOARD ---
ws = wb.create_sheet("Dashboard_Main")

df_old = df_summary[df_summary["Snapshot"]=="first"].groupby("Category")["Count"].sum().reset_index()
df_last = df_summary[df_summary["Snapshot"]=="last"].groupby("Category")["Count"].sum().reset_index()

# Write OLD table
ws.append(["Category (OLD)", "Count"])
for r in df_old.itertuples(index=False):
    ws.append(list(r))

# PIE 1
pie1 = PieChart()
pie1.title = "Total Categories - Snapshot FIRST"
labels = Reference(ws, min_col=1, min_row=2, max_row=1+len(df_old))
data = Reference(ws, min_col=2, min_row=1, max_row=1+len(df_old))
pie1.add_data(data, titles_from_data=True)
pie1.set_categories(labels)
ws.add_chart(pie1, "E2")

# Write LAST table
start = len(df_old) + 5
ws.cell(row=start-1, column=1, value="Category (LAST)")
ws.cell(row=start-1, column=2, value="Count")
for r in df_last.itertuples(index=False):
    ws.append(list(r))

# PIE 2
pie2 = PieChart()
pie2.title = "Total Categories - Snapshot LAST"
labels2 = Reference(ws, min_col=1, min_row=start+1, max_row=start+len(df_last))
data2   = Reference(ws, min_col=2, min_row=start,   max_row=start+len(df_last))
pie2.add_data(data2, titles_from_data=True)
pie2.set_categories(labels2)
ws.add_chart(pie2, "E20")

# --- DETAIL DASHBOARDS ---
projects = df_summary["Project"].unique()

for proj in projects:
    ws2 = wb.create_sheet(f"Dashboard_{proj}")

    df_p = df_summary[df_summary["Project"] == proj].copy()

    # Full category list
    categories = sorted(df_p["Category"].unique())

    pivot = (
        df_p.pivot_table(
            index="Category",
            columns="Snapshot",
            values="Count",
            aggfunc="sum"
        )
        .reindex(categories)
        .fillna(0)
        .reset_index()
    )

    # Ensure first/last exist
    if "first" not in pivot.columns:
        pivot["first"] = 0
    if "last" not in pivot.columns:
        pivot["last"] = 0

    # Delta
    pivot["Delta"] = pivot["last"] - pivot["first"]

    # KPI box
    total_first = int(pivot["first"].sum())
    total_last = int(pivot["last"].sum())
    total_delta = total_last - total_first

    ws2["A1"] = "Project"
    ws2["B1"] = proj
    ws2["A2"] = "Total FIRST"
    ws2["B2"] = total_first
    ws2["A3"] = "Total LAST"
    ws2["B3"] = total_last
    ws2["A4"] = "Delta (LAST - FIRST)"
    ws2["B4"] = total_delta

    # Table header (row 6)
    ws2.append([])
    ws2.append(["Category", "first", "last", "Delta"])

    for row in pivot.itertuples(index=False):
        ws2.append(list(row))

    # Determine table size
    start_row = 7
    end_row = start_row + len(pivot) - 1

    # BAR CHART
    bar = BarChart()
    bar.type = "col"
    bar.title = f"{proj} – First vs Last Snapshot"
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Category"

    data_ref = Reference(ws2, min_col=2, max_col=3, min_row=start_row, max_row=end_row)
    cats_ref = Reference(ws2, min_col=1, min_row=start_row+1, max_row=end_row)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    ws2.add_chart(bar, "G2")

    # RADAR CHART
    radar = RadarChart()
    radar.title = f"{proj} – Radar (First vs Last)"
    radar.style = 26

    r_data = Reference(ws2, min_col=2, max_col=3, min_row=start_row, max_row=end_row)
    r_cats = Reference(ws2, min_col=1, min_row=start_row+1, max_row=end_row)
    radar.add_data(r_data, titles_from_data=True)
    radar.set_categories(r_cats)
    ws2.add_chart(radar, "G18")

wb.save(OUTPUT)
print("DONE →", OUTPUT)
